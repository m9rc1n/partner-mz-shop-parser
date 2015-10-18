from urllib.request import urlopen

from bs4 import BeautifulSoup
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import re

NONE_VALUE = 'None value'
ROOT_WEBSITE = 'http://www.biuroserwis.org/index.php?act=sklep&page='
ROOT_CATEGORY_NAME = 'Gandolf'
XLS_FILENAME = 'gandolf_'
XLS_DESC = 'Opis'
XLS_CODE = 'Kod'
XLS_NAME = 'Nazwa'


class Product:
    name = NONE_VALUE
    desc = NONE_VALUE
    code = NONE_VALUE
    title = NONE_VALUE

    def __init__(self):
        super().__init__()
        self.name = NONE_VALUE
        self.desc = NONE_VALUE
        self.code = NONE_VALUE
        self.title = NONE_VALUE

    def print(self):
        if self.name:
            print('    Found product ---------------------------- ')
            print('    Product Name: ', self.name)
            print('    Product Desc: ', self.desc)
            print('    Product Code: ', self.code)
            print('    Product title: ', self.title)

    def is_defined(self):
        if self.name and self.desc and self.code and self.title:
            return True
        else:
            return False


class Category:

    def __init__(self, name):
        super().__init__()
        self.name = name
        self.products = []

    products = []
    name = NONE_VALUE

    def add(self, product):
        self.products.append(product)


def parse_category(soup, depth, category_name):

    for i in range(1, 504):

        soup = BeautifulSoup(urlopen(ROOT_WEBSITE + str(i)).read())
        categories_list = soup('table', {"class": "itemTable"})
        replace_with_newline(soup, 'br')

        for category in categories_list[0].find_all('div', {'class': 'row'}):
            product = Product()
            product.name = category.find("div", {"class": "row_nazwa"}).text
            product.desc = category.find("div", {"class": "row_indeks"}).text.split('\n')
            # print(product.desc[1][9:])
            products.append(product)


def parse_code(soup, product):
    code_soup = soup('p', {'id': 'product_reference'})
    if code_soup:
        code = code_soup[0].span.text
        product.code = code


def replace_with_newline(soup, element):
    for e in soup.findAll(element):
        e.replace_with('\n')


def parse_desc(soup, product):
    # replace_with_newline(soup, 'p')
    desc_soup = soup('div', {'id': 'idTab1'})
    if desc_soup:
        desc = desc_soup[0].text
        product.desc = desc


def parse_name(soup, product):
    name_soup = soup('div', {'id': 'primary_block'})
    if name_soup:
        name = name_soup[0].h1.text
        product.name = name


def parse_product(soup):
    replace_with_newline(soup, 'br')

    product = Product()
    parse_name(soup, product)
    parse_code(soup, product)
    parse_desc(soup, product)
    if product.is_defined():
        product.print()
        # products.append(product)
    return product


def parse_products(soup, name):

    category = Category(name)

    for link in soup('p', {'class': 'product_desc'}):
        title = link.a.get('title')

        if title is not None:
            url_to_open = link.a.get('href')
            product_soup = BeautifulSoup(urlopen(url_to_open).read())
            product = parse_product(product_soup)
            product.title = title
            category.add(product)

    categories.append(category)


def write_product_to_xls_sheet(index, product, sheet):
    sheet.cell(row=index + 2, column=1).value = product.name
    sheet.cell(row=index + 2, column=2).value = product.desc[0]
    sheet.cell(row=index + 2, column=3).value = product.desc[1][9:]


def write_category_to_xls_book(book, category):
    sheet = book.create_sheet()
    sheet.title = "dsadas"
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Kod1"
    sheet.cell(row=1, column=3).value = "Kod2"
    return sheet


def write_to_xls():

    book = Workbook()

    sheet = write_category_to_xls_book(book, "aaa")

    for index, product in enumerate(products):
        write_product_to_xls_sheet(index, product, sheet)

    book.save('meritus.xlsx')

# --------------------------------------------------------------------------------------

categories = []
products_g = []
products_m = []

# soup = BeautifulSoup(urlopen(ROOT_WEBSITE).read())
# parse_category(soup, 0, ROOT_CATEGORY_NAME)
# write_to_xls()

# -------------------------------

gandolf = load_workbook(filename='gandolf.xlsx')
meritus = load_workbook(filename='meritus.xlsx')
s_gandolf = gandolf.active
s_meritus = meritus.active

j = 1

for r_gandolf in s_gandolf.rows:
    product = Product()
    product.name = r_gandolf[0].value  #.replace(" ", "")
    products_g.append(product)

for r_meritus in s_meritus.rows:
    product = Product()
    product.name = r_meritus[0].value  #.replace(" ", "")
    product.code = r_meritus[1].value
    product.desc = r_meritus[2].value
    products_m.append(product)

i = 1
for g in products_g:
    max_ratio = 0
    for m in products_m:
        ratio = difflib.SequenceMatcher(None, g.name.replace(" ", ""), m.name.replace(" ", "")).ratio()
        if ratio > 0.7 and ratio > max_ratio:  # g.name == m.name:
            s_gandolf.cell(row=i, column=2).value = m.code
            s_gandolf.cell(row=i, column=3).value = m.desc
            max_ratio = ratio
    i += 1

gandolf.save('gandolf_jeden_arkusz.xlsx')

# for r_meritus in s_meritus.rows:
#         print(r_meritus[0].value)
#         if j is 7000:
#             j = 0
#             continue
#         j += 1
        # if r_gandolf[0].value is r_meritus[0].value:
        #     print("Znaleziono")
        #     print(r_gandolf[0].value)

# book.save('gandolf_jeden_arkusz.xlsx')

# ----------------------------------

# wb = load_workbook(filename='new2.xlsx')
#
# sheet = wb.active
#
# for row in sheet.rows:
#     if row[9].value is not None:
#         row[9].value = "<br />".join(row[9].value.split("\n"))
#         # row[9].value = re.sub(r'\s+', ' ', row[2].value)
#         # row[9].value = re.sub(r'\s+', ' ', row[2].value)
#
# wb.save('new3.xlsx')