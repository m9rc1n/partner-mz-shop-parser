# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

import hashlib
import urllib.parse
import urllib.request as urllib2
import json


NR_KATALOGOWY = 2

url = 'http://93.157.103.46:42395/'

m = hashlib.md5()
m.update("vEkumUCuhucrufes7ebranadepre$wuy".encode('utf-8'))
m.update("79.96.198.212".encode('utf-8'))
key = m.hexdigest()

key_ = {'key': key}
params = urllib.parse.urlencode(key_)

list_of_types = 'api/list_of_types'
list_products_params = 'api/list_products_params/{0}/'
list_of_products = 'api/list_of_products/{0}/{1}/'
product_infou = 'api/product_info/{0}/{1}/'

types = json.loads(urllib2.urlopen(url + list_of_types, params.encode('utf-8')).read().decode())

book = Workbook()
sheet = book.get_active_sheet()

i = 1
bookNr = 0


def write_express():
    global ekspres, key, i, bookNr
    ekspres = name + "EKSPRES | " + naklad + cechy
    k = hashlib.md5()
    k.update(ekspres.encode('utf-8'))
    key = k.hexdigest()
    i += 1

    sheet.cell(row=i, column=NR_KATALOGOWY).value = key
    sheet.cell(row=i, column=14).value = 211
    sheet.cell(row=i, column=15).value = "Partner"
    sheet.cell(row=i, column=20).value = product_info["cena_ekspres"]
    sheet.cell(row=i, column=21).value = product_info["cena_ekspres"]
    sheet.cell(row=i, column=24).value = product_info["cena_ekspres"] * 1.23
    sheet.cell(row=i, column=25).value = product_info["cena_ekspres"] * 1.23
    sheet.cell(row=i, column=30).value = 0.23
    sheet.cell(row=i, column=45).value = ekspres
    sheet.cell(row=i, column=46).value = key
    termin = product_info["termin_ekspres"].split('-')
    opis = ekspres.replace(' | ', '<br />') + "<br />Termin dostawy: Od " + termin[0] + " do " + \
           termin[1] + " dni"
    sheet.cell(row=i, column=47).value = opis
    sheet.cell(row=i, column=48).value = opis
    sheet.cell(row=i, column=49).value = opis
    sheet.cell(row=i, column=63).value = "REKLAMA I POLIGRAFIA/Drukarnia"
    if i % 999 is 0:
        book.save(str(bookNr) + ".xlsx")
        bookNr += 1
        i = 1


for index1, one_type in enumerate(types):
    print(index1)
    url_product_params = url + list_products_params.format(one_type['id'])
    products_params = json.loads(urllib2.urlopen(url_product_params, params.encode('utf-8')).read().decode())
    if "atalog" in products_params['name']:
        continue
    for index2, group in enumerate(products_params['configs']['grupa']):
        url_list_of_products = url + list_of_products.format(one_type['id'], group)
        products = json.loads(urllib2.urlopen(url_list_of_products, params.encode('utf-8')).read().decode())
        for index3, product in enumerate(products):

            url_product_info = url + product_infou.format(one_type['id'], product['product'])
            product_info = json.loads(urllib2.urlopen(url_product_info, params.encode('utf-8')).read().decode())
            name = product_info['name'] + " | "
            naklad = "Nakład: " + product_info['naklad']
            cechy = ''
            for cecha in sorted(product_info['cechy']):
                cechy += " | " + cecha + ": "
                cechy += product_info['cechy'][cecha]

            standard = name + "STANDARD | " + naklad + cechy
            k = hashlib.md5()
            k.update(standard.encode('utf-8'))
            key = k.hexdigest()
            i += 1
            sheet.cell(row=i, column=NR_KATALOGOWY).value = key
            sheet.cell(row=i, column=14).value = 211
            sheet.cell(row=i, column=15).value = "Partner"
            sheet.cell(row=i, column=20).value = product_info["cena_standard"]
            sheet.cell(row=i, column=21).value = product_info["cena_standard"]
            sheet.cell(row=i, column=24).value = product_info["cena_standard"] * 1.23
            sheet.cell(row=i, column=25).value = product_info["cena_standard"] * 1.23
            sheet.cell(row=i, column=30).value = 0.23
            sheet.cell(row=i, column=45).value = standard
            sheet.cell(row=i, column=46).value = key
            termin = product_info["termin_standard"].split('-')
            opis = standard.replace(' | ', '<br />') + "<br />Termin dostawy: Od " + termin[0] + " do " + \
                   termin[1] + " dni"
            sheet.cell(row=i, column=47).value = opis
            sheet.cell(row=i, column=48).value = opis
            sheet.cell(row=i, column=49).value = opis
            sheet.cell(row=i, column=63).value = "REKLAMA I POLIGRAFIA/Drukarnia"

            if i % 999 is 0:
                book.save(str(bookNr) + ".xlsx")
                bookNr += 1
                i = 1
            if product_info["cena_ekspres"] is not 0:
                write_express()

            # if i > 500:
            #     break
                #
                # if i > 16:
                # break

        # if i > 500:
        #     break

    print(index1)
    print(i)